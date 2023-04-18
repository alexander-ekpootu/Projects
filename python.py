# encoding: utf-8
# coder:    Alexander Ekpo-Otu
# version:  3.0
# Edit To Do Multiple Files at a time and Output to the same file.
import wx
import os
import sys
import threading
import openpyxl
#!/usr/bin/env python

class LinkButton(wx.StaticText):
    def __init__(self, *args, **kwargs):
        wx.StaticText.__init__(self, *args, **kwargs)
        self.SetForegroundColour('blue')
        self.SetCursor(wx.Cursor(wx.CURSOR_HAND))
        self.SetFont(self.GetFont().Underlined())
        self.SetToolTip(self.GetLabel())

if getattr(sys, 'frozen', False):
    # If the script is run as a bundled executable, use the executable's directory
    script_dir = os.path.dirname(sys.executable)
else:
    # If the script is run from the terminal, use the current working directory
    script_dir = os.path.dirname(os.path.abspath(__file__))

os.chdir(script_dir)

if os.path.exists('settings'):
    load_sets = eval(open('settings', encoding='utf-8').read())
else:
    load_sets = {'is_maximized' : False, 'app_pos' : None, 'app_size' : (560, 400), 'ent_1' : '', 'ent_2' : ''}

eng = wx.App()

locale = wx.Locale(wx.LANGUAGE_ENGLISH)

fra = wx.Frame(parent=None, title='Experiment Data Organizer v2.0')
fra.SetIcon(wx.Icon(wx.Bitmap('images/app_icon.png')))  # Change the file format to '.png'

main_pan = wx.Panel(fra)
main_sizer = wx.BoxSizer(wx.VERTICAL)
main_pan.SetSizer(main_sizer)
# --------------------------------------------------------------------
head_pan = wx.Panel(main_pan)
head_pan.SetBackgroundColour('white')
head_pan.SetMinSize((-1, 70))
head_sizer = wx.BoxSizer(wx.HORIZONTAL)
head_pan.SetSizer(head_sizer)

main_sizer.Add(head_pan, proportion=0, flag=wx.EXPAND)

head_icon = wx.StaticBitmap(head_pan, -1, wx.Bitmap('images/head_icon.png'))
head_tit = wx.StaticText(head_pan, label=' Experiment Data Organizer v2.0')
head_tit.SetFont(wx.Font(12, wx.FONTFAMILY_DEFAULT, wx.NORMAL, wx.NORMAL, False, 'Verdana'))

head_sizer.AddStretchSpacer(1)
head_sizer.Add(head_icon, 0, flag=wx.EXPAND)
head_sizer.Add(head_tit, 0, flag=wx.ALIGN_CENTER)
head_sizer.AddStretchSpacer(1)

main_sizer.Add(wx.StaticLine(main_pan), flag=wx.EXPAND)
# --------------------------------------------------------------------

fix_wid_w = wx.StaticText(main_pan, label='Output Directory')
ext_fix_wid = fix_wid_w.GetSize()[0]
fix_wid_w.Destroy()

src_file_pan = wx.Panel(main_pan)
src_file_sizer = wx.BoxSizer(wx.HORIZONTAL)
src_file_pan.SetSizer(src_file_sizer)

main_sizer.AddSpacer(10)
main_sizer.Add(src_file_pan, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=3)

def open_src_file_folder(event):
    src_file_dir = src_file_ent.GetValue()

    if not os.path.exists(src_file_dir):
        wx.MessageDialog(fra, 'Make sure source file is specified and it exists.',
                         style=wx.ICON_INFORMATION | wx.OK).ShowModal()

    else:
        os.startfile(src_file_dir)


src_file_lbl = LinkButton(src_file_pan, label='Source File', size=(ext_fix_wid, -1), style=wx.ALIGN_RIGHT)
src_file_lbl.Bind(wx.EVT_LEFT_DOWN, open_src_file_folder)
src_file_sizer.Add(src_file_lbl, 0, flag=wx.ALIGN_CENTER)
src_file_sizer.AddSpacer(3)

src_file_ent = wx.TextCtrl(src_file_pan, value=load_sets['ent_1'])
src_file_ent.SetHint('Enter source files separated by ";"...')
src_file_sizer.Add(src_file_ent, 1, flag=wx.ALIGN_CENTER)
src_file_sizer.AddSpacer(3)


def src_file_sel_com(event):
    dialog = wx.FileDialog(None, "Select source files", wildcard="Excel Files (.xlsx)|*.xlsx",
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE)
    dialog.ShowModal()
    paths = dialog.GetPaths()
    if paths:
        src_file_ent.SetValue(";".join(paths))



src_file_btn = wx.Button(src_file_pan, label='Select...', style=wx.BU_EXACTFIT)
src_file_btn.SetMinSize((src_file_btn.GetSize().Width + 20, src_file_btn.GetSize().Height))
src_file_btn.SetBitmap(wx.Bitmap('images/select_folder.png'))
src_file_btn.Bind(wx.EVT_BUTTON, src_file_sel_com)
src_file_sizer.Add(src_file_btn, 0)

# --------------------------------------------------------------------

out_fold_pan = wx.Panel(main_pan)
out_fold_sizer = wx.BoxSizer(wx.HORIZONTAL)
out_fold_pan.SetSizer(out_fold_sizer)

def open_out_fold_folder(event):
    out_fold_dir = out_fold_ent.GetValue()

    if not os.path.exists(out_fold_dir):
        wx.MessageDialog(fra, 'Make sure output directory is specified and it exists.',
                         style=wx.ICON_INFORMATION | wx.OK).ShowModal()

    else:
        os.startfile(out_fold_dir)


out_fold_lbl = LinkButton(out_fold_pan, label='Output Directory', size=(ext_fix_wid, -1), style=wx.ALIGN_RIGHT)
out_fold_lbl.Bind(wx.EVT_LEFT_DOWN, open_out_fold_folder)
out_fold_sizer.Add(out_fold_lbl, 0, flag=wx.ALIGN_CENTER)
out_fold_sizer.AddSpacer(3)

out_fold_ent = wx.TextCtrl(out_fold_pan, value=load_sets['ent_2'])
out_fold_ent.SetHint('Enter output directory...')
out_fold_sizer.Add(out_fold_ent, 1, flag=wx.ALIGN_CENTER)
out_fold_sizer.AddSpacer(3)


def out_fold_sel_com(event):
    dialog = wx.DirDialog(None, "Choose output directory", "", wx.DD_DEFAULT_STYLE | wx.DD_DIR_MUST_EXIST)
    dialog.ShowModal()
    path = dialog.GetPath()
    if path:
        out_fold_ent.SetValue(path)


out_fold_btn = wx.Button(out_fold_pan, label='Select...', style=wx.BU_EXACTFIT)
out_fold_btn.SetMinSize((out_fold_btn.GetSize().Width + 20, out_fold_btn.GetSize().Height))
out_fold_btn.SetBitmap(wx.Bitmap('images/select_folder.png'))
out_fold_btn.Bind(wx.EVT_BUTTON, out_fold_sel_com)
out_fold_sizer.Add(out_fold_btn, 0)

main_sizer.AddSpacer(5)
main_sizer.Add(out_fold_pan, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=3)

main_sizer.AddSpacer(5)

# --------------------------------------------------------------------

fix_wid_w = wx.StaticText(main_pan, label='Start Process')
st_btn_w = fix_wid_w.GetSize()[0]
fix_wid_w.Destroy()

def del_st_btn_com(event):
    inp_files = src_file_ent.GetValue().split(';')
    out_dir = out_fold_ent.GetValue()
    mess = ''

    for inp_file in inp_files:
        if not os.path.exists(inp_file):
            mess = f'Make sure source files are specified and they exist.\nFile: {inp_file}'
            break
    if not os.path.exists(out_dir):
        mess = 'Make sure output directory is specified and it exists.'

    if mess:
        wx.MessageDialog(fra, mess, style=wx.ICON_INFORMATION | wx.OK).ShowModal()
    else:
        def pro_exec():
                st_btn.SetLabel('Working...')
                st_btn.Disable()

                for inp_file in inp_files:
                    inp_file_op = openpyxl.load_workbook(inp_file)
                    inp_sht = inp_file_op.active
                    inp_file_op.close()

                    inp_data = []

                    for row in inp_sht:
                        inp_data.append([c.value for c in row])

                    col_data = {}
                    col_data_d = {}
                    col_data_l = {}
                    
                    def process_d_section(rows):
                        processed_data = []
                        for row in rows:
                            if isinstance(row[0], str) and ':' in row[0]:
                                processed_data.extend(row[1:])
                            else:
                                processed_data.extend(row)
                        return processed_data

                    for r_i, row in enumerate(inp_data):
                        if row[0] == 'Subject:':
                            if row[1].startswith('JMR'):
                                sub = row[1]
                            else:
                                sub = row[2]
                            col_data[sub] = []
                            col_data_d[sub] = []
                            col_data_l[sub] = []

                        elif row[0] == 'H:':
                            steps = 0
                            for nxt_rows in inp_data[r_i:]:
                                if nxt_rows[0] == 'I:': break
                                steps += 1

                            for eve in inp_data[r_i + 1:r_i + steps]:
                                col_data[sub].extend(eve)

                        elif row[0] == 'D:':
                            steps = 0
                            for nxt_rows in inp_data[r_i:]:
                                if nxt_rows[0] == 'E:': break
                                steps += 1

                            for eve in inp_data[r_i + 1:r_i + steps]:
                                if isinstance(eve[0], str) and ':' in eve[0]:
                                    col_data_d[sub].extend(eve[1:])
                                else:
                                    col_data_d[sub].extend(eve)

                        elif row[0] == 'L:':
                            steps = 0
                            for nxt_rows in inp_data[r_i:]:
                                if nxt_rows[0] == 'M:' or steps > 9: break
                                steps += 1

                            for eve in inp_data[r_i + 1:r_i + steps]:
                                col_data_l[sub].extend(eve[1:][:8])

                    # --------------------------------------------------------------------

                    fin_out_data = [['C1', 'preCS1', 'preCS2', 'CS1', 'CS2', 'CS3', 'CS4', 'postCS1', 'postCS2']]

                    for sub in col_data:
                        fin_out_data.append([])
                        fin_out_data.append([sub])

                        trials = col_data[sub]
                        points = [eve for eve in trials if isinstance(eve, (float, int))]

                        t_vals = list(zip(*[iter(points)] * 8))
                        for t_i, t in enumerate(t_vals, start=1):
                            fin_out_data.append(['tr' + str(t_i), *t])

                    out_wb = openpyxl.Workbook()
                    out_sht_H = out_wb.active
                    out_sht_H.title = 'H Data'

                    for row in fin_out_data:
                        out_sht_H.append(row)


                    # --------------------------------------------------------------------

                    fin_out_data = [['C1', 'tr1', 'tr2', 'tr3', 'tr4', 'tr5', 'tr6', 'tr7', 'tr8', 'tr9', 'tr10', 'tr11', 'tr12']]

                    for trial in col_data_d:
                        data_points = col_data_d[trial]
                        data_points = [eve for eve in data_points if isinstance(eve, (float, int))]
                        del data_points[0]
                        fin_out_data.append([trial, *data_points])

                    out_sht_D = out_wb.create_sheet('D Data')

                    for row in fin_out_data:
                        out_sht_D.append(row)


                    # --------------------------------------------------------------------

                    fin_out_data = [['C1', 'tr1', 'tr2', 'tr3', 'tr4', 'tr5', 'tr6', 'tr7', 'tr8', 'tr9', 'tr10', 'tr11', 'tr12']]

                    for trial in col_data_l:
                        data_points = col_data_l[trial]
                        data_points = [eve for eve in data_points if isinstance(eve, (float, int))]
                        del data_points[0]
                        fin_out_data.append([trial, *data_points])

                    out_sht_L = out_wb.create_sheet('L Data')

                    for row in fin_out_data:
                        out_sht_L.append(row)


                    # --------------------------------------------------------------------
                    out_wb.save(os.path.join(out_dir, f'{os.path.splitext(os.path.basename(inp_file))[0]}_Output.xlsx'))

                st_btn.SetLabel('Start Process')
                st_btn.Enable()

                wx.MessageDialog(fra, 'Process completed successfully.', style=wx.ICON_INFORMATION).ShowModal()

        t = threading.Thread(target=pro_exec, daemon=True)
        t.start()
        t.join()


st_btn = wx.Button(main_pan, label='Start Process', size=(st_btn_w + 30, -1))
st_btn.SetBitmap(wx.Bitmap('images/btn_start.png'))
st_btn.Bind(wx.EVT_BUTTON, del_st_btn_com)

main_sizer.Add(st_btn, 0, flag=wx.ALIGN_CENTER)
main_sizer.AddSpacer(10)

# --------------------------------------------------------------------

if load_sets['is_maximized']:
    fra.Maximize()
else:
    fra.SetSize(load_sets['app_size'])
    if load_sets['app_pos']:
        fra.Move(load_sets['app_pos'])
    else:
        fra.CenterOnScreen()

def on_close(event):
    open('settings', 'w', encoding='utf-8').write(str({
        'is_maximized' : fra.IsMaximized(),
        'app_pos' : fra.GetPosition(),
        'app_size' : fra.GetSize(),
        'ent_1' : src_file_ent.GetValue(),
        'ent_2' : out_fold_ent.GetValue()
    }))

    event.Skip()

fra.Bind(wx.EVT_CLOSE, on_close)
fra.Show()
eng.MainLoop()
